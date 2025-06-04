import os
from pathlib import Path
from django.core.files.storage import default_storage
from django.conf import settings
from openpyxl import load_workbook
import tempfile


def get_excel_template(template_name):
    """
    Load an Excel template from storage (Minio) or local filesystem.
    
    Args:
        template_name (str): Name of the template file (e.g., 'plantilla.xlsx')
        
    Returns:
        openpyxl.Workbook: Loaded workbook object
        
    Raises:
        FileNotFoundError: If template doesn't exist in storage or locally
        Exception: If there's an error loading the workbook
    """
    # Storage path for the template
    storage_path = f'excel_templates/{template_name}'
    
    # First try to load from storage (Minio)
    if default_storage.exists(storage_path):
        try:
            # Open file from storage
            with default_storage.open(storage_path, 'rb') as storage_file:
                # Create a temporary file to work with openpyxl
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                    temp_file.write(storage_file.read())
                    temp_file_path = temp_file.name
                
                # Load workbook from temporary file
                workbook = load_workbook(temp_file_path)
                
                # Clean up temporary file
                os.unlink(temp_file_path)
                
                return workbook
                
        except Exception as e:
            print(f"Error loading template from storage: {e}")
            # Fall back to local file system
            pass
    
    # Fallback: try to load from local filesystem
    local_path = Path(settings.BASE_DIR) / 'excel_templates' / template_name
    
    if local_path.exists():
        try:
            return load_workbook(str(local_path))
        except Exception as e:
            raise Exception(f"Error loading local template {local_path}: {e}")
    
    # If neither storage nor local file exists
    raise FileNotFoundError(f"Template '{template_name}' not found in storage or local filesystem")


def save_workbook_to_response(workbook, filename):
    """
    Save a workbook to a temporary file and return the file path.
    Useful for creating downloads or further processing.
    
    Args:
        workbook (openpyxl.Workbook): The workbook to save
        filename (str): Desired filename
        
    Returns:
        str: Path to the temporary file
    """
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Save workbook to temporary file
    workbook.save(temp_file.name)
    
    return temp_file.name


def list_available_templates():
    """
    List all Excel templates available in storage and locally.
    
    Returns:
        dict: Dictionary with 'storage' and 'local' keys containing lists of template names
    """
    templates = {'storage': [], 'local': []}
    
    # Check storage templates
    try:
        # This is a simplified approach - in a real scenario you might need to 
        # implement a more sophisticated way to list files in storage
        storage_templates = ['plantilla.xlsx', 'plantilla_original.xlsx']
        for template in storage_templates:
            storage_path = f'excel_templates/{template}'
            if default_storage.exists(storage_path):
                templates['storage'].append(template)
    except Exception as e:
        print(f"Error listing storage templates: {e}")
    
    # Check local templates
    local_templates_dir = Path(settings.BASE_DIR) / 'excel_templates'
    if local_templates_dir.exists():
        for excel_file in local_templates_dir.glob('*.xlsx'):
            templates['local'].append(excel_file.name)
        for excel_file in local_templates_dir.glob('*.xls'):
            templates['local'].append(excel_file.name)
    
    return templates


def ensure_template_in_storage(template_name):
    """
    Ensure a template exists in storage. If not, try to upload it from local filesystem.
    
    Args:
        template_name (str): Name of the template file
        
    Returns:
        bool: True if template is available in storage, False otherwise
    """
    storage_path = f'excel_templates/{template_name}'
    
    # Check if already in storage
    if default_storage.exists(storage_path):
        return True
    
    # Try to upload from local
    local_path = Path(settings.BASE_DIR) / 'excel_templates' / template_name
    
    if local_path.exists():
        try:
            from django.core.files.base import ContentFile
            
            with open(local_path, 'rb') as local_file:
                content = local_file.read()
                content_file = ContentFile(content)
                default_storage.save(storage_path, content_file)
                
            return default_storage.exists(storage_path)
            
        except Exception as e:
            print(f"Error uploading template to storage: {e}")
            return False
    
    return False